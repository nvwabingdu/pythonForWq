import tkinter as tk
from datetime import datetime
import os
import pandas as pd
import tkinter.messagebox
from PIL import ImageGrab, Image
import io
import threading
import time

COLUMNS = [
    "登记时间",
    "订单号",
    "客户账号",
    "SKU",
    "补发原因",
    "姓名",
    "电话",
    "收货地址",
    "补发型号",
    "数量",
    "登记人",
    "凭证图片",
]

REQUIRED_ORDER = ["订单号", "客户账号", "SKU", "补发原因", "姓名", "电话", "收货地址", "补发型号","数量","登记人",]

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("补发登记")
        self.attributes("-topmost", True)

        self._cursor = 0
        self._last_clip = None
        self.clip_image = None
        self.waiting_image = False

        right = tk.Frame(self)
        right.pack(fill="both", expand=True, padx=0, pady=0)

        widths = {
            "登记时间": 140,
            "订单号": 110,
            "客户账号": 110,
            "SKU": 90,
            "补发原因": 130,
            "姓名": 75,
            "电话": 105,
            "收货地址": 240,
            "补发型号": 130,
            "数量": 65,
            "登记人": 85,
            "凭证图片": 100,
        }

        self.vars = {}
        font = ("Microsoft YaHei", 9)

        header = tk.Frame(right)
        header.pack(fill="x")
        for col in COLUMNS:
            w = widths.get(col, 100)
            lbl = tk.Label(header, text=col, bd=1, relief="solid", font=font,
                           width=max(6, int(w/10)), anchor="center")
            lbl.pack(side="left", padx=1, pady=0)

        inp = tk.Frame(right)
        inp.pack(fill="x", pady=0)

        self.reg_time_var = tk.StringVar(value=self.get_today_cn_date())
        self.user_var = tk.StringVar()

        for col in COLUMNS:
            w = widths.get(col, 100)
            if col == "登记时间":
                v = self.reg_time_var
                e = tk.Entry(inp, textvariable=v, font=font, width=max(6, int(w/10)),
                             bd=1, relief="solid", justify="center", state="readonly")
            elif col == "凭证图片":
                v = tk.StringVar(value="无")
                e = tk.Entry(inp, textvariable=v, font=font, width=max(6, int(w/10)),
                             bd=1, relief="solid", justify="center", state="readonly")
                self.vars[col] = v
            else:
                v = tk.StringVar()
                e = tk.Entry(inp, textvariable=v, font=font, width=max(6, int(w/10)),
                             bd=1, relief="solid", justify="center")
                self.vars[col] = v
            e.pack(side="left", padx=1, pady=0)

        self.btn_img = tk.Button(inp, text="图片", command=self.start_wait_image,
                            font=font, width=10,
                            bg="#4285F4", fg="white",
                            bd=1, relief="solid")
        self.btn_img.pack(side="left", padx=1, pady=0)

        btn_gen = tk.Button(inp, text="生成", command=self.on_generate,
                            font=font, width=6,
                            bg="#28a845", fg="white",
                            bd=1, relief="solid")
        btn_gen.pack(side="left", padx=1, pady=0)

        self.start_listening = False
        self.after(200, self.poll_clipboard)

        self.update()
        self.geometry(f"{self.winfo_width()}x{self.winfo_height()}")
        self.resizable(False, False)

    def get_today_cn_date(self):
        return datetime.now().strftime("%Y年%m月%d日")

    def set_current_column(self, text):
        if self._cursor >= len(REQUIRED_ORDER):
            return
        col = REQUIRED_ORDER[self._cursor]
        self.vars[col].set(text)
        self._cursor += 1

    def poll_clipboard(self):
        try:
            clip = self.clipboard_get().strip()
        except tk.TclError:
            clip = ""

        if clip and clip != self._last_clip:
            if not self.start_listening:
                self.start_listening = True
                self._last_clip = clip
            else:
                self._last_clip = clip
                self.set_current_column(clip)

        self.after(150, self.poll_clipboard)

    def start_wait_image(self):
        self.clear_clipboard()
        self.clip_image = None
        self.vars["凭证图片"].set("等待粘贴...")
        self.btn_img.config(text="等待粘贴图片", state="disabled", bg="#FF9800")
        self.waiting_image = True
        threading.Thread(target=self.listen_clipboard_image, daemon=True).start()

    def clear_clipboard(self):
        try:
            self.clipboard_clear()
            self.clipboard_append("")
            self.update()
        except:
            pass

    def listen_clipboard_image(self):
        while self.waiting_image:
            time.sleep(0.3)
            try:
                img = ImageGrab.grabclipboard()
                if isinstance(img, Image.Image):
                    self.clip_image = img
                    self.waiting_image = False
                    self.after(0, self.on_image_caught)
                    break
            except:
                continue

    def on_image_caught(self):
        self.vars["凭证图片"].set("已获取图片")
        self.btn_img.config(text="图片", state="normal", bg="#4285F4")

    def on_generate(self):
        missing = [c for c in REQUIRED_ORDER if not self.vars[c].get().strip()]
        if missing:
            tkinter.messagebox.showwarning("提示", f"缺少：{', '.join(missing)}")
            return

        record = {
            "登记时间": self.reg_time_var.get(),
            "订单号": self.vars["订单号"].get().strip(),
            "客户账号": self.vars["客户账号"].get().strip(),
            "SKU": self.vars["SKU"].get().strip(),
            "补发原因": self.vars["补发原因"].get().strip(),
            "姓名": self.vars["姓名"].get().strip(),
            "电话": self.vars["电话"].get().strip(),
            "收货地址": self.vars["收货地址"].get().strip(),
            "补发型号": self.vars["补发型号"].get().strip(),
            "数量": self.vars["数量"].get().strip(),
            "登记人": self.vars["登记人"].get().strip(),
            "凭证图片": "已上传",
        }

        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        file = os.path.join(desktop, f"{datetime.now():%Y-%m-%d}_补发登记.xlsx")

        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage

        if os.path.exists(file):
            wb = load_workbook(file)
            ws = wb.active
            row_num = ws.max_row + 1
        else:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(COLUMNS)
            row_num = 2

        row_data = [record[c] for c in COLUMNS]
        ws.append(row_data)

        if self.clip_image is not None:
            # ========== 固定 2cm × 2cm 纯手工计算（无依赖，绝对不报错） ==========
            # 固定尺寸：2 厘米 × 2 厘米
            CELL_CM = 2.0

            # Excel 标准单位换算（通用所有版本）
            col_width = CELL_CM * 7.82        # 列宽 = 2cm
            row_height = CELL_CM * 27.68      # 行高 = 2cm
            img_px = int(CELL_CM * 96 / 2.54) # 图片像素 = 2cm

            # 设置单元格
            ws.column_dimensions['L'].width = col_width
            ws.row_dimensions[row_num].height = row_height

            # 插入图片（缩放到 2cm，原图保留）
            img_byte_arr = io.BytesIO()
            self.clip_image.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)

            img = XLImage(img_byte_arr)
            img.width = img_px
            img.height = img_px
            ws.add_image(img, f"L{row_num}")
            # ==================================================================

        wb.save(file)

        # 重置
        self._cursor = 0
        self._last_clip = None
        self.start_listening = False
        self.user_var.set("")
        self.clip_image = None
        self.waiting_image = False
        for c in self.vars:
            self.vars[c].set("" if c != "登记时间" else self.get_today_cn_date())
        self.btn_img.config(text="图片", state="normal", bg="#4285F4")

    def reset_for_next(self):
        self._cursor = 0
        self._last_clip = None
        self.reg_time_var.set(self.get_today_cn_date())
        for col in REQUIRED_ORDER:
            self.vars[col].set("")
        self.attributes("-topmost", True)

if __name__ == "__main__":
    app = App()
    app.mainloop()