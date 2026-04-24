import tkinter as tk
from datetime import datetime
import os
import pandas as pd
import tkinter.messagebox
from PIL import ImageGrab, Image
import io
import threading
import time
import openpyxl
import pyperclip
import pyautogui
import win32clipboard

# ===================== 固定配置 =====================
COLUMNS = [
    "登记时间", "订单号", "客户账号", "SKU", "补发原因",
    "姓名", "电话", "收货地址", "补发型号", "数量", "登记人", "凭证图片",
]

REQUIRED_ORDER = [
    "订单号", "客户账号", "SKU", "补发原因", "姓名", "电话",
    "收货地址", "补发型号", "数量", "登记人",
]

# =============== 钉钉发送线程（全局中断标记）===============
stop_signal = False
send_thread = None

# =========================================================
# ===================== 主程序窗口 ======================
# =========================================================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("补发登记 + 自动发钉钉")
        self.attributes("-topmost", True)

        self._cursor = 0
        self._last_clip = None
        self.clip_image = None
        self.waiting_image = False

        # 界面布局
        right = tk.Frame(self)
        right.pack(fill="both", expand=True, padx=0, pady=0)

        widths = {
            "登记时间":140,"订单号":110,"客户账号":110,"SKU":90,"补发原因":130,
            "姓名":75,"电话":105,"收货地址":240,"补发型号":130,"数量":65,
            "登记人":85,"凭证图片":100,
        }

        self.vars = {}
        font = ("Microsoft YaHei", 9)

        # 表头
        header = tk.Frame(right)
        header.pack(fill="x")
        for col in COLUMNS:
            w = widths.get(col,100)
            lbl = tk.Label(header, text=col, bd=1, relief="solid", font=font,
                           width=max(6,int(w/10)), anchor="center")
            lbl.pack(side="left", padx=1, pady=0)

        # 输入行
        inp = tk.Frame(right)
        inp.pack(fill="x", pady=0)

        self.reg_time_var = tk.StringVar(value=self.get_today_cn_date())
        self.user_var = tk.StringVar()

        for col in COLUMNS:
            w = widths.get(col,100)
            if col == "登记时间":
                v = self.reg_time_var
                e = tk.Entry(inp, textvariable=v, font=font, width=max(6,int(w/10)),
                             bd=1, relief="solid", justify="center", state="readonly")
            elif col == "凭证图片":
                v = tk.StringVar(value="无")
                e = tk.Entry(inp, textvariable=v, font=font, width=max(6,int(w/10)),
                             bd=1, relief="solid", justify="center", state="readonly")
                self.vars[col] = v
            else:
                v = tk.StringVar()
                e = tk.Entry(inp, textvariable=v, font=font, width=max(6,int(w/10)),
                             bd=1, relief="solid", justify="center")
                self.vars[col] = v
            e.pack(side="left", padx=1, pady=0)

        # 按钮：图片
        self.btn_img = tk.Button(inp, text="图片", command=self.start_wait_image,
                            font=font, width=10, bg="#4285F4", fg="white", bd=1, relief="solid")
        self.btn_img.pack(side="left", padx=1, pady=0)

        # 按钮：生成
        btn_gen = tk.Button(inp, text="生成", command=self.on_generate,
                            font=font, width=6, bg="#28a845", fg="white", bd=1, relief="solid")
        btn_gen.pack(side="left", padx=1, pady=0)

        # ===================== 【发到钉钉】按钮（核心新增）=====================
        self.btn_ding = tk.Button(inp, text="发到钉钉", command=self.toggle_dingtalk,
                            font=font, width=10, bg="#FF4444", fg="white", bd=1, relief="solid")
        self.btn_ding.pack(side="left", padx=1, pady=0)

        # 剪贴板监听
        # 添加置空
        # 只向剪贴板写入一个点，不粘贴、不模拟按键、不弹窗
        pyperclip.copy("默认值")
        self.start_listening = False
        self.after(200, self.poll_clipboard)
        self.update()
        self.geometry(f"{self.winfo_width()}x{self.winfo_height()}")
        self.resizable(False, False)

    # ===================== 工具函数 =====================
    def get_today_cn_date(self):
        return datetime.now().strftime("%Y年%m月%d日")


    def set_current_column(self, text):
        if self._cursor >= len(REQUIRED_ORDER):
            return
        # 遍历所有必填字段，按顺序找第一个为空的
        for col in REQUIRED_ORDER:
            # 判断当前字段是否为空
            if not self.vars[col].get().strip():
                self.vars[col].set(text)  # 为空就赋值
                break  # 找到第一个就停止，不继续往下填
        self._cursor += 1

    def poll_clipboard(self):
        try:
            clip = self.clipboard_get().strip()
        except tk.TclError:
            clip = ""

        if clip and clip != self._last_clip:
            if not self.start_listening:
                # 添加置空
                pyperclip.copy("默认值")
                self.start_listening = True
                self._last_clip = clip
            else:
                self._last_clip = clip
                self.set_current_column(clip)
        self.after(150, self.poll_clipboard)

    def start_wait_image(self):
        self.clear_clipboard()
        self.clip_image = None
        self.vars["凭证图片"].set("等待粘贴...")
        self.btn_img.config(text="等待粘贴图片", state="disabled", bg="#FF9800")
        self.waiting_image = True
        threading.Thread(target=self.listen_clipboard_image, daemon=True).start()

    def clear_clipboard(self):
        try:
            self.clipboard_clear()
            self.clipboard_append("")
            self.update()
        except:
            pass

    def listen_clipboard_image(self):
        while self.waiting_image:
            time.sleep(0.3)
            try:
                img = ImageGrab.grabclipboard()
                if isinstance(img, Image.Image):
                    self.clip_image = img
                    self.waiting_image = False
                    self.after(0, self.on_image_caught)
                    break
            except:
                continue

    def on_image_caught(self):
        self.vars["凭证图片"].set("已获取图片")
        self.btn_img.config(text="图片", state="normal", bg="#4285F4")

    # ===================== 生成Excel =====================
    def on_generate(self):
        missing = [c for c in REQUIRED_ORDER if not self.vars[c].get().strip()]
        if missing:
            tkinter.messagebox.showwarning("提示", f"缺少：{', '.join(missing)}")
            return

        record = {
            "登记时间": self.reg_time_var.get(),
            "订单号": self.vars["订单号"].get().strip(),
            "客户账号": self.vars["客户账号"].get().strip(),
            "SKU": self.vars["SKU"].get().strip(),
            "补发原因": self.vars["补发原因"].get().strip(),
            "姓名": self.vars["姓名"].get().strip(),
            "电话": self.vars["电话"].get().strip(),
            "收货地址": self.vars["收货地址"].get().strip(),
            "补发型号": self.vars["补发型号"].get().strip(),
            "数量": self.vars["数量"].get().strip(),
            "登记人": self.vars["登记人"].get().strip(),
            "凭证图片": "",
        }

        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        file = os.path.join(desktop, f"补发登记.xlsx")

        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage

        if os.path.exists(file):
            wb = load_workbook(file)
            ws = wb.active
            row_num = ws.max_row + 1
        else:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(COLUMNS)
            row_num = 2

        row_data = [record[c] for c in COLUMNS]
        ws.append(row_data)

        from openpyxl.styles import numbers
        for col in range(1, len(COLUMNS)+1):
            cell = ws.cell(row=row_num, column=col)
            cell.number_format = numbers.FORMAT_TEXT

        if self.clip_image is not None:
            CELL_CM = 1.0
            col_width = CELL_CM * 7.82
            row_height = CELL_CM * 27.68
            img_px = int(CELL_CM * 96 / 2.54)

            ws.column_dimensions['L'].width = col_width
            ws.row_dimensions[row_num].height = row_height

            img_byte_arr = io.BytesIO()
            self.clip_image.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            img = XLImage(img_byte_arr)
            img.width = img_px
            img.height = img_px
            ws.add_image(img, f"L{row_num}")

        wb.save(file)

        self._cursor = 0
        self._last_clip = None
        # 添加置空
        pyperclip.copy("默认值")
        self.start_listening = False
        self.user_var.set("")
        self.clip_image = None
        self.waiting_image = False
        for c in self.vars:
            self.vars[c].set("" if c != "登记时间" else self.get_today_cn_date())
        self.btn_img.config(text="图片", state="normal", bg="#4285F4")

    def reset_for_next(self):
        self._cursor = 0
        self._last_clip = None
        self.reg_time_var.set(self.get_today_cn_date())
        for col in REQUIRED_ORDER:
            self.vars[col].set("")
        self.attributes("-topmost", True)

    # ===================== 【发到钉钉】启动 / 中断 =====================
    def toggle_dingtalk(self):
        global stop_signal, send_thread

        if send_thread is None or not send_thread.is_alive():
            # 启动
            stop_signal = False
            self.btn_ding.config(text="停止发送", bg="#000000", fg="white")
            send_thread = threading.Thread(target=self.run_send_dingtalk, daemon=True)
            send_thread.start()
        else:
            # 强制中断
            stop_signal = True
            self.btn_ding.config(text="发到钉钉", bg="#FF4444")
            tkinter.messagebox.showinfo("中断", "已强制停止发送！")

    # ===================== 钉钉发送主逻辑 =====================
    def run_send_dingtalk(self):
        global stop_signal
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        excel_path = os.path.join(desktop, f"补发登记.xlsx")

        if not os.path.exists(excel_path):
            tkinter.messagebox.showerror("错误", "未找到今天的补发登记表格！")
            self.after(0, lambda: self.btn_ding.config(text="发到钉钉", bg="#FF4444"))
            return

        time.sleep(2)

        # 打开Excel
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        img_map = self.get_all_images_in_sheet(ws)
        end_row = ws.max_row

        for row in range(2, end_row + 1):
            if stop_signal:
                break

            print(f"正在发送第 {row} 行")
            pyautogui.hotkey("ctrl", "a")
            pyautogui.press("backspace")
            time.sleep(0.15)

            last_col = ws.max_column
            for col in range(1, last_col + 1):
                if stop_signal:
                    break

                pos = (row, col)
                cell_val = ws.cell(row=row, column=col).value

                # 图片
                if pos in img_map:
                    img = img_map[pos]
                    image = Image.open(io.BytesIO(img._data()))
                    self.copy_image_to_clipboard(image)
                    pyautogui.hotkey("ctrl", "v")
                    time.sleep(3)
                else:
                    # 文字
                    val = str(cell_val).strip() if cell_val else ""
                    if val:
                        pyperclip.copy(val)
                        pyautogui.hotkey("ctrl", "v")

                # 换行（非最后一列）
                if col != last_col:
                    pyautogui.hotkey("ctrl", "enter")

            # 发送
            if not stop_signal:
                pyautogui.press("enter")

        # 结束
        self.after(0, lambda: self.btn_ding.config(text="发到钉钉", bg="#FF4444"))
        if stop_signal:
            tkinter.messagebox.showinfo("停止", "已手动停止")
        else:
            tkinter.messagebox.showinfo("完成", "全部发送完毕！")

    def get_all_images_in_sheet(self, ws):
        image_map = {}
        for img in ws._images:
            r = img.anchor._from.row + 1
            c = img.anchor._from.col + 1
            image_map[(r, c)] = img
        return image_map

    def copy_image_to_clipboard(self, img):
        output = io.BytesIO()
        img.convert("RGB").save(output, "BMP")
        data = output.getvalue()[14:]
        output.close()
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
        win32clipboard.CloseClipboard()

if __name__ == "__main__":
    app = App()
    app.mainloop()