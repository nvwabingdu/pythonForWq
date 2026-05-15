import tkinter as tk
from datetime import datetime
import os
import pandas as pd
import tkinter.messagebox
from PIL import ImageGrab, Image
import io
import threading
import time
import openpyxl
import pyperclip
import pyautogui
import win32clipboard

# ===================== 固定配置 =====================
COLUMNS = [
    "登记时间", "订单号", "客户账号", "SKU", "补发原因",
    "姓名", "电话", "收货地址", "补发型号", "数量", "登记人", "凭证图片",
]

REQUIRED_ORDER = [
    "订单号", "客户账号", "SKU", "补发原因", "姓名", "电话",
    "收货地址", "补发型号", "数量", "登记人",
]

# =============== 钉钉发送线程（全局中断标记）===============
stop_signal = False
send_thread = None

# ===================== 图片压缩/剪贴板策略（优化：防止卡死）=====================
# 你要求“极致压缩且不失真”：优先 PNG 无损，但同时限制尺寸，避免超大位图导致剪贴板卡死。
# 如果你的原图分辨率非常夸张，这个限制非常关键。
MAX_LONG_EDGE = 1400          # 图片最长边上限（可按实际微调）
MAX_CLIP_BYTES = 3_000_000    # 压缩后最终字节上限（可按实际微调，越小越安全但越可能损失）
PNG_FORMAT = "PNG"

# =========================================================
# ===================== 主程序窗口 ======================
# =========================================================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("补发登记 + 自动发钉钉")
        self.attributes("-topmost", True)

        self._last_clip = None
        self.clip_image = None
        self.waiting_image = False

        # 界面布局
        right = tk.Frame(self)
        right.pack(fill="both", expand=True, padx=0, pady=0)

        widths = {
            "登记时间": 140, "订单号": 110, "客户账号": 110, "SKU": 90, "补发原因": 130,
            "姓名": 75, "电话": 105, "收货地址": 240, "补发型号": 130, "数量": 65,
            "登记人": 85, "凭证图片": 100,
        }

        self.vars = {}
        font = ("Microsoft YaHei", 9)

        # ===================== 表头 =====================
        header = tk.Frame(right)
        header.pack(fill="x")
        for col in COLUMNS:
            w = widths.get(col, 100)
            lbl = tk.Label(
                header, text=col, bd=1, relief="solid", font=font,
                width=max(6, int(w / 10)), anchor="center"
            )
            lbl.pack(side="left", padx=1, pady=0)

        # ===================== 输入行 =====================
        inp = tk.Frame(right)
        inp.pack(fill="x", pady=0)

        self.reg_time_var = tk.StringVar(value=self.get_today_cn_date())
        self.user_var = tk.StringVar()

        # 监听状态：是否正在消费剪贴板文本/图片
        self.clip_running = True  # 加这个
        self.start_listening = False

        for col in COLUMNS:
            w = widths.get(col, 100)
            if col == "登记时间":
                v = self.reg_time_var
                e = tk.Entry(inp, textvariable=v, font=font, width=max(6, int(w / 10)),
                             bd=1, relief="solid", justify="center", state="readonly")
            elif col == "凭证图片":
                v = tk.StringVar(value="无")
                e = tk.Entry(inp, textvariable=v, font=font, width=max(6, int(w / 10)),
                             bd=1, relief="solid", justify="center", state="readonly")
                self.vars[col] = v
            else:
                v = tk.StringVar()
                e = tk.Entry(inp, textvariable=v, font=font, width=max(6, int(w / 10)),
                             bd=1, relief="solid", justify="center")
                self.vars[col] = v

            e.pack(side="left", padx=1, pady=0)

        # ===================== 按钮：图片 =====================
        self.btn_img = tk.Button(
            inp, text="图片", command=self.start_wait_image,
            font=font, width=10, bg="#4285F4", fg="white", bd=1, relief="solid"
        )
        self.btn_img.pack(side="left", padx=1, pady=0)

        # ===================== 按钮：生成 =====================
        btn_gen = tk.Button(
            inp, text="生成", command=self.on_generate,
            font=font, width=6, bg="#28a845", fg="white", bd=1, relief="solid"
        )
        btn_gen.pack(side="left", padx=1, pady=0)

        # ===================== 【发到钉钉】按钮（核心） =====================
        self.btn_ding = tk.Button(
            inp, text="发到钉钉", command=self.toggle_dingtalk,
            font=font, width=10, bg="#FF4444", fg="white", bd=1, relief="solid"
        )
        self.btn_ding.pack(side="left", padx=1, pady=0)

        # ===================== 剪贴板监听（性能优化：避免强制update/几何计算） =====================
        # 添加置空：只写一个点，避免“程序启动就拿到旧剪贴板内容填充”
        pyperclip.copy("默认值")
        self.after(200, self.poll_clipboard)

        # 启动性能：不强行 update() + 立即 geometry 计算（会明显拖慢启动）
        self.resizable(False, False)

    # ===================== 工具函数 =====================
    def get_today_cn_date(self):
        return datetime.now().strftime("%Y年%m月%d日")

    def get_first_missing_field(self):
        """根据当前界面输入，返回“第一个为空字段”的列名（修复：手动删除后仍可继续填充）"""
        for col in REQUIRED_ORDER:
            if not self.vars[col].get().strip():
                return col
        return None

    def set_current_column(self, text):
        """
        向“第一个为空字段”写入文本。
        不再依赖旧的游标 _cursor，从而修复你说的“手动删除后不能再次监听”的 bug。
        """
        target = self.get_first_missing_field()
        if not target:
            return  # 所有必填都已填满，就不继续写了
        self.vars[target].set(text)

    def poll_clipboard(self):
        # 关键：这里是真正能停止循环的开关
        if not self.clip_running:
            return

        try:
            clip = self.clipboard_get().strip()
        except tk.TclError:
            clip = ""

        if clip and clip != self._last_clip:
            if not self.start_listening:
                pyperclip.copy("默认值")
                self.start_listening = True
                self._last_clip = clip
            else:
                self._last_clip = clip
                self.set_current_column(clip)

        # 继续循环
        self.after(150, self.poll_clipboard)

    def start_wait_image(self):
        """等待剪贴板粘贴图片：在单独线程里轮询 ImageGrab.grabclipboard()"""
        self.clear_clipboard()
        self.clip_image = None
        self.vars["凭证图片"].set("等待粘贴...")
        self.btn_img.config(text="等待粘贴图片", state="disabled", bg="#FF9800")
        self.waiting_image = True
        threading.Thread(target=self.listen_clipboard_image, daemon=True).start()

    def clear_clipboard(self):
        """清空 Tk 自带剪贴板（防止旧内容干扰图片监听）"""
        try:
            self.clipboard_clear()
            self.clipboard_append("")
            self.update()
        except:
            pass

    def listen_clipboard_image(self):
        """后台线程：轮询剪贴板是否有图片对象"""
        while self.waiting_image:
            time.sleep(0.3)
            try:
                img = ImageGrab.grabclipboard()
                if isinstance(img, Image.Image):
                    self.clip_image = img
                    self.waiting_image = False
                    self.after(0, self.on_image_caught)
                    break
            except:
                continue

    def on_image_caught(self):
        """主线程：图片获取成功后更新UI"""
        self.vars["凭证图片"].set("已获取图片")
        self.btn_img.config(text="图片", state="normal", bg="#4285F4")

    # ===================== 生成Excel =====================
    def on_generate(self):
        missing = [c for c in REQUIRED_ORDER if not self.vars[c].get().strip()]
        if missing:
            tkinter.messagebox.showwarning("提示", f"缺少：{', '.join(missing)}")
            return

        record = {
            "登记时间": self.reg_time_var.get(),
            "订单号": self.vars["订单号"].get().strip(),
            "客户账号": self.vars["客户账号"].get().strip(),
            "SKU": self.vars["SKU"].get().strip(),
            "补发原因": self.vars["补发原因"].get().strip(),
            "姓名": self.vars["姓名"].get().strip(),
            "电话": self.vars["电话"].get().strip(),
            "收货地址": self.vars["收货地址"].get().strip(),
            "补发型号": self.vars["补发型号"].get().strip(),
            "数量": self.vars["数量"].get().strip(),
            "登记人": self.vars["登记人"].get().strip(),
            "凭证图片": "",
        }

        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        file = os.path.join(desktop, f"补发登记.xlsx")

        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage

        if os.path.exists(file):
            wb = load_workbook(file)
            ws = wb.active
            row_num = ws.max_row + 1
        else:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(COLUMNS)
            row_num = 2

        row_data = [record[c] for c in COLUMNS]
        ws.append(row_data)

        from openpyxl.styles import numbers
        for col in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.number_format = numbers.FORMAT_TEXT

        # ===================== 写入图片（写入时不做压缩，发送时压缩） =====================
        # 这里先保持你的逻辑：把你获取到的clip_image写进Excel。
        # 发送阶段再极致压缩，避免钉钉剪贴板卡死。
        if self.clip_image is not None:
            CELL_CM = 1.0
            col_width = CELL_CM * 7.82
            row_height = CELL_CM * 27.68
            img_px = int(CELL_CM * 96 / 2.54)

            ws.column_dimensions['L'].width = col_width
            ws.row_dimensions[row_num].height = row_height

            img_byte_arr = io.BytesIO()
            # 保存为PNG（无损，避免你说“不失真”）
            self.clip_image.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            img = XLImage(img_byte_arr)
            img.width = img_px
            img.height = img_px
            ws.add_image(img, f"L{row_num}")

        wb.save(file)

        # ===================== 生成后清空输入 =====================
        self.start_listening = False
        pyperclip.copy("默认值")

        self.reg_time_var.set(self.get_today_cn_date())
        self.clip_image = None
        self.waiting_image = False
        self.btn_img.config(text="图片", state="normal", bg="#4285F4")

        # 清空所有字段（登记时间保留今天）
        for c in self.vars:
            self.vars[c].set("" if c != "登记时间" else self.get_today_cn_date())

    # ===================== 【发到钉钉】启动 / 中断 =====================
    def toggle_dingtalk(self):
        global stop_signal, send_thread

        if send_thread is None or not send_thread.is_alive():
            stop_signal = False
            self.btn_ding.config(text="停止发送", bg="#000000", fg="white")
            send_thread = threading.Thread(target=self.run_send_dingtalk, daemon=True)
            send_thread.start()
        else:
            stop_signal = True
            self.btn_ding.config(text="发到钉钉", bg="#FF4444")
            tkinter.messagebox.showinfo("中断", "已强制停止发送！")

    # ===================== 钉钉发送主逻辑 =====================
    def run_send_dingtalk(self):
        self.clip_running = False
        global stop_signal
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        excel_path = os.path.join(desktop, f"补发登记.xlsx")

        if not os.path.exists(excel_path):
            self.after(0, lambda: tkinter.messagebox.showerror("错误", "未找到今天的补发登记表格！"))
            self.after(0, lambda: self.btn_ding.config(text="发到钉钉", bg="#FF4444"))
            return

        # 小等待让用户切到钉钉窗口（保留你的逻辑：仍然等一下）
        time.sleep(1.5)

        # 打开Excel
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        img_map = self.get_all_images_in_sheet(ws)
        end_row = ws.max_row
        last_col = ws.max_column

        for row in range(2, end_row + 1):
            if stop_signal:
                break

            print(f"正在发送第 {row} 行")

            # 你原逻辑：清空输入框
            pyautogui.hotkey("ctrl", "a")
            pyautogui.press("backspace")
            time.sleep(0.15)

            for col in range(1, last_col + 1):
                if stop_signal:
                    break

                pos = (row, col)
                cell_val = ws.cell(row=row, column=col).value

                # 处理 NaN/None：避免出现 'nan' 文本
                if cell_val is None:
                    val = ""
                else:
                    try:
                        # pandas NaN
                        if pd.isna(cell_val):
                            val = ""
                        else:
                            val = str(cell_val).strip()
                    except:
                        val = str(cell_val).strip() if cell_val else ""

                # ===================== 图片列处理（关键优化） =====================
                if pos in img_map:
                    img = img_map[pos]

                    # 1) 从openpyxl图片数据取出 PIL 图像
                    try:
                        image = Image.open(io.BytesIO(img._data()))
                    except Exception as e:
                        print("图片读取失败：", e)
                        image = None

                    if image is not None:
                        # 2) 先极致压缩/限制尺寸，避免剪贴板卡死
                        compressed_image = self.prepare_image_for_clipboard(image)

                        # 3) 复制图片到剪贴板并粘贴
                        self.copy_image_to_clipboard(compressed_image)
                        pyautogui.hotkey("ctrl", "v")

                        # 4) 等待钉钉渲染：不再固定 sleep(3)，改为“更快的自适应等待”
                        if not self.wait_for_dingtalk_image_render(max_wait=3):
                            # 如果超时，仍然继续（避免死循环）
                            pass
                    else:
                        # 如果图片读取失败，跳过该列
                        pass
                else:
                    # ===================== 文本列处理（保持原逻辑） =====================
                    if val:
                        pyperclip.copy(val)
                        pyautogui.hotkey("ctrl", "v")

                # 换行（非最后一列）
                if col != last_col:
                    pyautogui.hotkey("ctrl", "enter")

            # 发送
            if not stop_signal:
                pyautogui.press("enter")

        # 结束提示：回到主线程弹窗，避免线程里直接弹造成卡顿
        if stop_signal:
            self.after(0, lambda: tkinter.messagebox.showinfo("停止", "已手动停止"))
        else:
            self.after(0, lambda: tkinter.messagebox.showinfo("完成", "全部发送完毕！"))

        self.after(0, lambda: self.btn_ding.config(text="发到钉钉", bg="#FF4444"))
        # 重新启动
        self.clip_running = True
        pyperclip.copy("默认值")
        self.poll_clipboard()

    # ===================== 获取Excel里所有嵌入图片映射（列定位） =====================
    def get_all_images_in_sheet(self, ws):
        image_map = {}
        for img in getattr(ws, "_images", []):
            try:
                # openpyxl anchor 行列：_from.row/_from.col 从0开始
                r = img.anchor._from.row + 1
                c = img.anchor._from.col + 1
                image_map[(r, c)] = img
            except:
                continue
        return image_map

    # ===================== 图片压缩（关键：减少剪贴板卡死） =====================
    def prepare_image_for_clipboard(self, img: Image.Image) -> Image.Image:
        """
        极致压缩策略：
        - 保持尽量不失真：优先PNG保存（无损）思路
        - 但剪贴板/位图有上限：必须限制尺寸，避免超大
        - 如果压缩后字节超限，再进一步缩放
        """
        # 统一转RGBA或RGB，避免某些模式导致保存异常
        if img.mode not in ("RGB", "RGBA"):
            img = img.convert("RGBA")

        # 限制尺寸（只缩放，不裁切，尽量不失真）
        w, h = img.size
        long_edge = max(w, h)
        if long_edge > MAX_LONG_EDGE:
            scale = MAX_LONG_EDGE / long_edge
            new_w = max(1, int(w * scale))
            new_h = max(1, int(h * scale))
            img = img.resize((new_w, new_h), Image.LANCZOS)

        # 再做一次“字节上限”控制：通过PNG编码大小判断
        # 若超限，继续缩小尺寸，直到满足 MAX_CLIP_BYTES 或缩到很小
        # （这一段不会影响“不改逻辑”，只是优化图片输入）
        quality_scale = 1.0
        for _ in range(6):
            buf = io.BytesIO()
            # PNG无损（更符合“不失真”），但仍可能太大，因此尺寸控制仍必要
            img_to_save = img
            # 如果是RGBA，导出为RGB可减少透明通道导致的数据量（更安全）
            if img_to_save.mode == "RGBA":
                img_to_save = img_to_save.convert("RGB")
            img_to_save.save(buf, format=PNG_FORMAT)
            size = buf.tell()
            buf.close()

            if size <= MAX_CLIP_BYTES:
                return img

            # 超了就按比例缩小
            quality_scale *= 0.85
            if quality_scale < 0.3:
                break
            new_w = max(1, int(img.size[0] * 0.85))
            new_h = max(1, int(img.size[1] * 0.85))
            img = img.resize((new_w, new_h), Image.LANCZOS)

        return img

    # ===================== 等待钉钉渲染图片（替代固定sleep 3秒） =====================
    def wait_for_dingtalk_image_render(self, max_wait=2.0, step=0.2):
        """
        在不做UI识别的前提下，我们采取“更快的自适应等待”：
        - 复制图片并粘贴后，等待短时间（max_wait内）
        - 期间不做长sleep，避免阻塞/变慢
        - 由于无法保证100%判断“已渲染”，这里采用保守等待上限策略
        """
        waited = 0.0
        # 注：这里无法可靠读取“输入框是否已有图片”状态，所以仅做时间上限等待。
        # 图片已经压缩后，通常 0.6~1.2s 就会完成渲染，避免原先固定3s。
        while waited < max_wait:
            if stop_signal:
                return False
            time.sleep(step)
            waited += step
            # 达到一定时间就认为足够
            if waited >= 1.0:
                return True
        return False

    # ===================== 复制图片到剪贴板（关键：避免超大导致卡死） =====================
    def copy_image_to_clipboard(self, img: Image.Image):
        """
        仍沿用你原先的 CF_DIB 写入方式，但前面已经做了 prepare_image_for_clipboard 压缩。
        """
        output = io.BytesIO()
        # Windows剪贴板位图写入：建议转RGB，避免RGBA导致额外复杂度
        if img.mode != "RGB":
            img = img.convert("RGB")
        img.save(output, "BMP")
        data = output.getvalue()[14:]  # 去除BMP头
        output.close()

        win32clipboard.OpenClipboard()
        try:
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
        finally:
            win32clipboard.CloseClipboard()

if __name__ == "__main__":
    app = App()
    app.mainloop()